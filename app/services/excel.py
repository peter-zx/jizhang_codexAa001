import json
import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from app.models import Person, StoredFile, User
from app.services.ledger import normalize_period, normalize_status, valid_person_name
from app.services.storage import export_path

logger = logging.getLogger(__name__)

FIELD_DEFINITIONS = [
    ("serial_no", "序号"),
    ("name", "姓名"),
    ("age", "年龄"),
    ("gender", "性别"),
    ("contact_phone", "联系方式"),
    ("settlement_period", "结算所属期"),
    ("employment_status", "在职离职"),
    ("entry_date", "入职时间"),
    ("sfid", "SFid"),
    ("sfid_expires_at", "SFid过期时间"),
    ("disability_cert_id", "残疾证ID"),
    ("cert_issued_at", "发证时间"),
    ("household_address", "户籍住址"),
    ("household_type", "户籍性质"),
    ("work_area", "在职地区"),
    ("placement_period", "安置时间"),
    ("salary_card", "工资卡"),
    ("service_fee", "服务费"),
    ("payroll_type", "发薪类型"),
    ("gross_pay", "应发"),
    ("return_amount", "应返"),
    ("bank_card_id", "银行卡ID"),
    ("bank_name", "开户行"),
    ("emergency_contact", "紧急联系人"),
    ("emergency_phone", "紧急联系人电话"),
    ("emergency_relation", "紧急联系人关系"),
    ("education", "文化程度"),
    ("marital_status", "婚否"),
    ("disability_type1", "残疾类型1"),
    ("disability_level1", "残疾等级1"),
    ("disability_type2", "残疾类型2"),
    ("disability_level2", "残疾等级2"),
    ("disability_part", "残疾部位"),
    ("disability_reason", "残疾原因"),
    ("channel", "渠道"),
    ("note", "备注"),
    ("owner", "分销商"),
    ("monthly_confirmed", "本月核准"),
    ("confirmed_at", "核准时间"),
]
DEFAULT_EXPORT_FIELDS = ["name", "age", "disability_type1", "disability_level1", "disability_type2", "disability_level2"]
TEMPLATE_FIELDS = [
    "serial_no", "name", "work_area", "placement_period", "salary_card", "service_fee", "payroll_type", "gross_pay", "return_amount",
    "settlement_period", "employment_status", "note", "channel", "household_address", "household_type", "sfid",
    "sfid_expires_at", "disability_cert_id", "cert_issued_at", "contact_phone", "emergency_contact",
    "emergency_phone", "emergency_relation", "education", "marital_status", "bank_card_id", "bank_name",
    "disability_type1", "disability_level1", "disability_type2", "disability_level2", "entry_date", "age",
    "gender", "disability_part", "disability_reason",
]

FIELD_ALIASES = {
    "serial_no": ["序号"],
    "name": ["姓名"],
    "sfid": ["SFid", "sfid"],
    "sfid_expires_at": ["SFid过期时间"],
    "disability_cert_id": ["残疾证ID", "残疾证id", "残疾证号", "残疾人证号"],
    "cert_issued_at": ["发证时间"],
    "work_area": ["在职地区"],
    "placement_period": ["安置时间"],
    "salary_card": ["工资卡"],
    "service_fee": ["服务费"],
    "payroll_type": ["发薪类型"],
    "gross_pay": ["应发"],
    "return_amount": ["应返"],
    "settlement_period": ["结算所属期"],
    "employment_status": ["在职离职"],
    "note": ["备注"],
    "channel": ["渠道"],
    "household_address": ["户籍住址"],
    "household_type": ["户籍性质"],
    "contact_phone": ["联系方式"],
    "emergency_contact": ["紧急联系人"],
    "emergency_phone": ["紧急联系人电话"],
    "emergency_relation": ["紧急联系人关系"],
    "education": ["文化程度"],
    "marital_status": ["婚否"],
    "bank_card_id": ["银行卡ID"],
    "bank_name": ["开户行"],
    "disability_type1": ["残疾类型1", "残疾类型"],
    "disability_level1": ["残疾等级1", "残疾等级"],
    "disability_type2": ["残疾类型2"],
    "disability_level2": ["残疾等级2"],
    "entry_date": ["入职时间"],
    "age": ["年龄"],
    "gender": ["性别"],
    "disability_part": ["残疾部位"],
    "disability_reason": ["残疾原因"],
}


def _num(value) -> float:
    if value in (None, "", "-"):
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _header_indexes(headers: list[object]) -> dict[str, int]:
    normalized = {str(value).strip(): idx for idx, value in enumerate(headers) if value not in (None, "")}
    indexes: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                indexes[field] = normalized[alias]
                break
    return indexes


def import_people_from_xlsx(db: Session, path: Path, owner: User, operator: User) -> int:
    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        indexes = _header_indexes(headers)
        if "name" not in indexes:
            raise ValueError("表格缺少必填列：姓名")
        known_indexes = set(indexes.values())
        created = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            payload = {}
            for field, idx in indexes.items():
                payload[field] = row[idx] if idx < len(row) else None
            if not valid_person_name(payload.get("name")):
                continue
            payload["name"] = str(payload["name"]).strip()
            payload["sfid"] = str(payload.get("sfid") or "").strip() or None
            payload["disability_cert_id"] = str(payload.get("disability_cert_id") or "").strip() or None
            payload["service_fee"] = _num(payload.get("service_fee") if "service_fee" in payload else owner.default_service_fee)
            payload["gross_pay"] = _num(payload.get("gross_pay"))
            payload["return_amount"] = _num(payload.get("return_amount"))
            payload["settlement_period"] = normalize_period(payload.get("settlement_period"))
            payload["employment_status"] = normalize_status(payload.get("employment_status"))
            extra = {}
            for idx, header in enumerate(headers):
                if idx in known_indexes or header in (None, ""):
                    continue
                value = row[idx] if idx < len(row) else None
                if value not in (None, ""):
                    extra[str(header).strip()] = value
            payload["extra_fields"] = json.dumps(extra, ensure_ascii=False) if extra else None

            if payload["sfid"]:
                existing = db.query(Person).filter(Person.sfid == payload["sfid"]).first()
                if existing:
                    logger.info("跳过重复 SFid：%s", payload["sfid"])
                    continue
            if payload["disability_cert_id"]:
                existing = db.query(Person).filter(Person.disability_cert_id == payload["disability_cert_id"]).first()
                if existing:
                    logger.info("跳过重复残疾证ID：%s", payload["disability_cert_id"])
                    continue
            db.add(Person(owner_id=owner.id, **payload))
            created += 1
        db.commit()
        logger.info("用户 %s 为 %s 导入 %s 条人员记录", operator.username, owner.username, created)
        return created
    except Exception:
        db.rollback()
        logger.exception("导入 Excel 失败: %s", path)
        raise


def _export_value(person: Person, field: str):
    if field == "owner":
        return person.owner.display_name
    if field == "monthly_confirmed":
        return "已核准" if person.monthly_confirmed else "待核准"
    if field == "confirmed_at":
        return person.confirmed_at.strftime("%Y-%m-%d %H:%M") if person.confirmed_at else ""
    return getattr(person, field, "")


def export_people_to_xlsx(db: Session, people: list[Person], current_user: User, period: str, fields: list[str] | None = None) -> StoredFile:
    try:
        allowed = dict(FIELD_DEFINITIONS)
        selected = [field for field in (fields or DEFAULT_EXPORT_FIELDS) if field in allowed]
        if not selected:
            selected = DEFAULT_EXPORT_FIELDS

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "月度对账"
        sheet.append([allowed[field] for field in selected])
        for item in people:
            sheet.append([_export_value(item, field) for field in selected])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B87")
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 28)
        if sheet.max_row >= 2 and sheet.max_column >= 1:
            table_ref = f"A1:{sheet.cell(row=1, column=sheet.max_column).column_letter}{sheet.max_row}"
            table = Table(displayName="MonthlyLedger", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        sheet.freeze_panes = "A2"
        stored_name, path = export_path("最新数据")
        workbook.save(path)
        record = StoredFile(owner_id=current_user.id, file_type="export", original_name=None, stored_name=stored_name, path=str(path))
        db.add(record)
        db.commit()
        return record
    except Exception:
        db.rollback()
        logger.exception("导出 Excel 失败")
        raise


def create_template_xlsx() -> tuple[str, Path]:
    try:
        allowed = dict(FIELD_DEFINITIONS)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "人员导入模板"
        sheet.append([allowed[field] for field in TEMPLATE_FIELDS if field in allowed])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B87")
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 24)
        sheet.freeze_panes = "A2"
        stored_name, path = export_path("people_template")
        workbook.save(path)
        return stored_name, path
    except Exception:
        logger.exception("生成导入模板失败")
        raise
